"""
Quart: This program is a 'Demo' that will read patents and store them in a spreadsheet

"""

from selenium import webdriver
#from selenium.webdriver.common.keys import Keys
import chromedriver_autoinstaller
from bs4 import BeautifulSoup
import time
import requests 
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import readPDF as pdf


chromedriver_autoinstaller.install()
download_dir='C:\\Users\\Acer\\Downloads\\'
#Set options for chrome
options = Options()
profile = {
           "plugins.plugins_list": [{"enabled": False, "name": "Chrome PDF Viewer"}], # Disable Chrome's PDF Viewer
           "download.default_directory": download_dir , 
           "download.extensions_to_open": "applications/pdf",
           "plugins.always_open_pdf_externally": True
           }

options.add_experimental_option("prefs", profile)

browser=webdriver.Chrome(options=options)
#Start 13000
StartID=5000
EndID=14000
countExpedient=0

for i in range(StartID,EndID):
    #This iteration gets each file
    urlExp="https://vidoc.impi.gob.mx/visor?usr=SIGA&texp=SI&tdoc=E&id=MX/a/2015/00"+str(i)
    response= requests.get(urlExp)
    status= response.status_code
    if status==200:
        browser.get(urlExp)
        time.sleep(1)
        path=''
        pathPdf=''
        expedient_name=''
        exp_html=''
        table=''
        exp_html = BeautifulSoup(browser.page_source, 'lxml')
        time.sleep(3)
        table=exp_html.find('table')
        expedient_name=exp_html.find('h3').text
        expedient_name=expedient_name.replace(' ','_')  
        expedient_name=expedient_name.replace('/','_') 
        path=download_dir+expedient_name+'.xlsx'
        res=os.path.isfile(path)
        if res==False:
            if table is not None:
                #The complete table if exists
                table_rows = table.findAll('tr')
                #Iterate all the rows in the table
                countPatent=0
                for tr in table_rows:
                    #Every <td> has one element only
                    txt_number=''
                    txt_barcode=''
                    txt_document=''
                    txt_desc=''
                    txt_type=''
                    txt_date=''
                    pdf_name=''
                    pdf_file_name=''
                    if tr.nextSibling!='\n':
                        td = tr.findAll('td')
                        fieldPosition=0
                        for t in td:
                            fieldPosition=fieldPosition+1
                            #Id of the input for each row index 0
                            #MainContent_gdDoctosExpediente_ImageButton1_0 
                            btn=t.findChildren('input',recursive=True)
                            if btn:
                                chunks=str(btn[0]).split(' ')
                                #Getting ID alone
                                parts=chunks[2].split('=')
                                val_name=parts[1] 
                                javaScript = "document.getElementsByName("+val_name+")[0].click();"
                                browser.execute_script(javaScript)
                                #Get the name of the pdf document
                                time.sleep(3) #Time sleep to give time to read the 'modal-text'
                                pdf_name=txt_document.replace('/','_')
                                pdf_file_name=pdf_name+'.pdf'
                                pdf_source=''
                                pdf_source = browser.find_element_by_tag_name('iframe').get_attribute("src")
                                time.sleep(2)
                                if pdf_source!='':
                                    #Get the url of the source
                                    browser.get(pdf_source)
                                    time.sleep(5)
                                    #Finf the href with innerText 'aquí'
                                    lst_link=browser.find_elements_by_tag_name('a')
                                    for link in lst_link:   
                                        if link.text=='aquí':
                                            link.click()
                                            #Wait 'X' seconds for download
                                            time.sleep(10) 
                                            #Get the expedient web page again, due to change of pages
                                            #it is needed to come back to a prior page
                                            browser.execute_script('window.history.go(-1)')
                                            browser.refresh()
                                            #pathPdf=''
                                            #pathPdf=download_dir+pdf_file_name
                                            #resPdf=pdf.readPdf(pathPdf,'','txt')
                                            #if resPdf:
                                            #print('Pdf ready for:',txt_document)
                                            continue
                                            
                                           
                            else:
                                """
                                Structure of <tr/>:
                                Number (1),BarCode (2),Document (3),Description (4),Type (5),Date (6),PDF (7)
                                """  
                                if fieldPosition==1:
                                    txt_number=t.text
                                    continue
                                if fieldPosition==2:
                                    txt_barcode=t.text
                                    continue
                                if fieldPosition==3:
                                    txt_document=t.text
                                    continue 
                                if fieldPosition==4:
                                    txt_desc=t.text
                                    continue
                                if fieldPosition==5:
                                    txt_type=t.text
                                    continue  
                                if fieldPosition==6:
                                    txt_date=t.text 
                                    continue            
                        #End of loop of every td in a single row 
                        #Excel process
                        flag=False
                        flag=os.path.isfile(path)
                        if flag:
                            #Expedient xls already exists
                            wb = load_workbook(path)
                            ws = wb['Patents']
                         
                        else:
                            #Expedient xls is new 
                            wb = Workbook()
                            ws = wb.active
                            ws.title = "Patents" 
                            #Write(row,column)
                            #Headers (h1,...)
                            h1 = ws.cell(row = 1, column = 1)
                            h1.value = 'Number'
                            h2 = ws.cell(row = 1, column = 2)
                            h2.value = 'Bar code'
                            h3 = ws.cell(row = 1, column = 3)
                            h3.value = 'Document'
                            h4 = ws.cell(row = 1, column = 4)
                            h4.value = 'Description'
                            h5 = ws.cell(row = 1, column = 5)
                            h5.value = 'Type'
                            h6 = ws.cell(row = 1, column = 6)
                            h6.value = 'Date'
                            h7=ws.cell(row=1,column=7)
                            h7.value='Pdf file'
                    
                        #Row
                        countPatent=countPatent+1
                        #As rows start at 1, then only inside row values the counPatent is added 1 again
                        number = ws.cell(row = countPatent+1, column = 1)
                        number.value = txt_number
                        barCode=ws.cell(row = countPatent+1, column = 2)
                        barCode.value=txt_barcode
                        document=ws.cell(row = countPatent+1, column = 3)
                        document.value=txt_document
                        desc=ws.cell(row = countPatent+1, column = 4)
                        desc.value=txt_desc
                        vtype=ws.cell(row = countPatent+1, column = 5)
                        vtype.value=txt_type
                        vdate=ws.cell(row = countPatent+1, column = 6)
                        vdate.value=txt_date
                        vpdf=ws.cell(row = countPatent+1, column = 7)
                        vpdf.value=pdf_file_name                   
                        wb.save(path) 
                     
                   
                        #if countRow==1:
                        #    break
                #End of row loop 
                countExpedient=countExpedient+1
                print('Expedients so far:',str(countExpedient))                    
                if countExpedient==47:
                    break        
        
browser.quit() 

    
       
        
   