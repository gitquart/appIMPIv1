"""
Code to join all information from patents in workbook
into one main workbook
"""
import os
from openpyxl import Workbook
from openpyxl import load_workbook

dir='C:\\Users\\Acer\\Documents\\quart\\PatentDemo\\'
dir_wb_join='C:\\Users\\Acer\\Documents\\quart\\condensed.xlsx'

for x in os.listdir(dir):
    if x.endswith(".xlsx"):
        flag=False
        flag=os.path.isfile(dir_wb_join)
        if flag:  
            cwb = load_workbook(dir_wb_join)
            cws = cwb['Patents'] 
                
        else:
            cwb = Workbook()
            cws = cwb.active
            cws.title = "Patents" 
            #Write(row,column)
            #Headers (h1,...)
            h1 = cws.cell(row=1, column=1)
            h1.value = 'Number'
            h2 = cws.cell(row=1, column=2)
            h2.value = 'Bar code'
            h3 = cws.cell(row=1, column=3)
            h3.value = 'Document'
            h4 = cws.cell(row=1, column=4)
            h4.value = 'Description'
            h5 = cws.cell(row=1, column=5)
            h5.value = 'Type'
            h6 = cws.cell(row=1, column=6)
            h6.value = 'Date'
            h7 = cws.cell(row=1, column=7)
            h7.value = 'Pdf file'
            
        wb = load_workbook(dir+x)
        ws = wb['Patents']
            
        #This method gets the tuple of values needed
        for row in ws.iter_rows(min_row=2,values_only=True):
            cws.append(row)
            cwb.save(dir_wb_join)    
                
print('Done!')